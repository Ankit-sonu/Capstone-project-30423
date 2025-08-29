from openpyxl import load_workbook, Workbook

source_file = "rawdata.xlsx"
destination_file = "cleandata.xlsx"

# Load the source_file

workbook = load_workbook(source_file)
worksheet = workbook.active

#Read the header
rows = list(worksheet.iter_rows(values_only=True))
headers = rows[0]
data_rows = rows[1:]
cleaned_data = [headers]

for row in data_rows:
    eid, name, dept, salary = row
    if eid is None or name is None or dept is None:
        continue
    try:
        salary= float(salary)
    except (ValueError,TypeError):
        continue
    cleaned_data.append([eid, name, dept, salary])

newbook = Workbook()
newsheet = newbook.active
newsheet.title="Cleaned data"
for row in cleaned_data:
    newsheet.append(row)
newbook.save(destination_file)
print("Data is extracted cleaned and loaded...!")