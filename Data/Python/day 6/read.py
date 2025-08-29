from openpyxl import load_workbook

# Load the workbook in python

workbook = load_workbook(r"C:\Users\admin\Desktop\python codes\day 6")

#sheet = workbook.active
sheet = workbook['Sheet2']


first = True
for row in sheet.iter_rows(min_row=2, values_only=True):
    empid,name,dept, desg = row
    if first:
        first = False
        continue
    print(f" Empid : {empid} , Name : {name}")