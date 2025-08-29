from openpyxl import load_workbook
from employeedao import EmployeeDao
from employee import Employee

class EmployeeService:

    def _init_(self):
        self.dao = EmployeeDao()

    def clean_employee_data(self, employees):
        cleaned = []
        for e in employees:
            if isinstance(e.empid, int) and isinstance(e.name, str) and e.age > 18:
                cleaned.append((e.empid, e.name.strip(), int(e.age), e.dept, e.notes, float(e.salary), e.remarks))
        return cleaned

    def load_employee_data_todb(self, employees):
        self.dao.add_employee(employees)

    @staticmethod
    def read_from_excel(filename):
        try:
            wb = load_workbook(filename)
            sheet = wb.active
            employees = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                empid, name, age, dept, notes, salary, remarks = row
                employees.append(Employee(empid, name, age, dept, notes, salary, remarks))
            return employees
        except Exception as e:
            print(f"Error reading Excel: {e}")
            return []
